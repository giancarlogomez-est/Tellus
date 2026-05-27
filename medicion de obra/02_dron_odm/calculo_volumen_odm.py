# =============================================================================
# CÁLCULO DE VOLÚMENES — FLUJO DRON + OpenDroneMap
# =============================================================================
# Entrada:
#   dem_pre.tif            → DSM del vuelo de referencia (antes de la obra)
#   dem_post.tif           → DSM del vuelo de medición
#   eje_via.dwg / .dxf / .geojson  → Eje de vía de AutoCAD o QGIS
#
# Salida:
#   reporte_volumen_odm.xlsx       → Tabla por progresiva + resumen
#   diferencia_dem.tif             → Raster ΔZ (post − pre)
#   heatmap_volumenes.png          → Mapa de calor de ΔZ
#   superficies_civil3d.xml        → LandXML para importar en Civil 3D
# =============================================================================

import numpy as np
import os, sys

# ---------------------------------------------------------------------------
# PARÁMETROS — editar antes de cada vuelo
# ---------------------------------------------------------------------------
DEM_PRE           = "dem_pre.tif"        # ← DSM vuelo 1 (referencia)
DEM_POST          = "dem_post.tif"       # ← DSM vuelo 2 (medición)

# EJE DE LA VÍA — acepta .dwg, .dxf o .geojson
EJE_VIA           = "eje_via.dwg"        # ← archivo del eje (AutoCAD o GeoJSON)

# SISTEMA DE PROYECCIÓN
# Código EPSG del DWG/DXF. Ejemplos Colombia:
#   9377  → MAGNA-SIRGAS / Origen-Nacional (recomendado)
#   3116  → MAGNA-SIRGAS / Colombia Bogota zone
#   32618 → WGS 84 / UTM zona 18N
#   4326  → WGS 84 geográfico
EPSG_PROYECTO     = 9377                 # ← EPSG del sistema de coordenadas del proyecto

ESPACIADO_M       = 20.0                 # metros entre secciones transversales
ANCHO_CORREDOR_M  = 28.0                 # metros desde el eje hacia cada lado
PROG_INICIO_M     = 5300.0               # progresiva inicial (K5+300 → 5300)
FACT_ESPONJ       = 1.25
FACT_CONTRAC      = 0.90
PASO_MUESTREO     = 4                    # densidad malla LandXML (2=0.5m, 4=1m, 8=2m)
ARCHIVO_REPORTE   = "reporte_volumen_odm.xlsx"
ARCHIVO_DIFDEM    = "diferencia_dem.tif"
ARCHIVO_HEATMAP   = "heatmap_volumenes.png"
ARCHIVO_LANDXML   = "superficies_civil3d.xml"
# ---------------------------------------------------------------------------


def verificar_dependencias():
    faltantes = []
    for pkg in ["rasterio", "numpy", "geopandas", "shapely",
                "matplotlib", "openpyxl", "scipy", "ezdxf", "pyproj"]:
        try:
            __import__(pkg)
        except ImportError:
            faltantes.append(pkg)
    if faltantes:
        print("Instala las dependencias con:")
        print(f"  pip install {' '.join(faltantes)}")
        sys.exit(1)


# ---------------------------------------------------------------------------
# Carga del eje de vía (DWG / DXF / GeoJSON)
# ---------------------------------------------------------------------------
def cargar_eje(eje_path, epsg_proyecto):
    """
    Carga el eje de la vía desde .dwg, .dxf o .geojson.
    Si es DWG/DXF, usa convertir_eje_dwg.py para extraer la polilínea.
    Devuelve la ruta al GeoJSON listo para usar.
    """
    import geopandas as gpd

    ext = os.path.splitext(eje_path)[1].lower()

    if ext in (".dwg", ".dxf"):
        print(f"  Eje AutoCAD detectado ({ext.upper()}). Convirtiendo a GeoJSON...")
        # Importar funciones del convertidor
        conv_dir = os.path.dirname(os.path.abspath(__file__))
        sys.path.insert(0, conv_dir)
        from convertir_eje_dwg import (dwg_a_dxf, listar_polilineas,
                                        seleccionar_polilinea, extraer_coordenadas,
                                        exportar_geojson)

        # Convertir DWG → DXF si es necesario
        if ext == ".dwg":
            dxf_path = dwg_a_dxf(eje_path)
        else:
            dxf_path = eje_path

        # Extraer polilínea (modo auto: toma la más larga si hay varias)
        polilineas = listar_polilineas(dxf_path)
        if not polilineas:
            print("ERROR: No se encontraron polilíneas en el DXF.")
            sys.exit(1)

        # Seleccionar la polilínea más larga automáticamente
        entidad = max(polilineas, key=lambda p: p[3])[4]
        coords  = extraer_coordenadas(entidad)

        # Guardar GeoJSON junto al DWG
        geojson_path = os.path.splitext(eje_path)[0] + "_eje.geojson"
        exportar_geojson(coords, epsg_proyecto, geojson_path)
        print(f"  GeoJSON generado: {geojson_path}")
        return geojson_path

    elif ext == ".geojson" or ext == ".json":
        return eje_path
    else:
        # Intentar como shapefile u otro formato vectorial soportado por geopandas
        try:
            gdf = gpd.read_file(eje_path)
            geojson_path = os.path.splitext(eje_path)[0] + "_eje.geojson"
            gdf.to_file(geojson_path, driver="GeoJSON")
            return geojson_path
        except Exception as e:
            print(f"ERROR: Formato de eje no soportado ({ext}): {e}")
            sys.exit(1)


def alinear_dems(dem_pre_path, dem_post_path):
    """Reproyecta dem_post a la misma grilla que dem_pre."""
    import rasterio
    from rasterio.warp import reproject, Resampling

    with rasterio.open(dem_pre_path) as src_pre:
        perfil_pre = src_pre.profile.copy()
        datos_pre  = src_pre.read(1)
        nodata_pre = src_pre.nodata

    with rasterio.open(dem_post_path) as src_post:
        datos_post_alin = np.zeros_like(datos_pre, dtype=np.float32)
        nodata_post = src_post.nodata

        reproject(
            source=rasterio.band(src_post, 1),
            destination=datos_post_alin,
            src_transform=src_post.transform,
            src_crs=src_post.crs,
            dst_transform=src_pre.transform,
            dst_crs=src_pre.crs,
            resampling=Resampling.bilinear,
        )

    return datos_pre, datos_post_alin, perfil_pre, nodata_pre


def calcular_diferencia(datos_pre, datos_post, perfil, nodata, ruta_salida):
    """Genera raster ΔZ = post − pre y lo guarda."""
    import rasterio

    mask = (datos_pre == nodata) | (datos_post == 0)
    delta = np.where(mask, np.nan, datos_post.astype(float) - datos_pre.astype(float))

    perfil_out = perfil.copy()
    perfil_out.update(dtype="float32", nodata=np.nan)

    with rasterio.open(ruta_salida, "w", **perfil_out) as dst:
        dst.write(delta.astype(np.float32), 1)

    print(f"Raster ΔZ guardado: {ruta_salida}")
    return delta


def generar_heatmap(delta, perfil, ruta_salida):
    """Genera mapa de calor PNG del raster ΔZ."""
    import matplotlib.pyplot as plt
    import matplotlib.colors as mcolors
    import rasterio.transform as rtransform

    transform = perfil["transform"]
    ancho = delta.shape[1]
    alto  = delta.shape[0]
    x0, y0 = transform.c, transform.f
    x1 = x0 + transform.a * ancho
    y1 = y0 + transform.e * alto
    extent = [x0, x1, y1, y0]

    vmax = np.nanpercentile(np.abs(delta), 95)
    vmax = max(vmax, 0.1)

    fig, ax = plt.subplots(figsize=(14, 8))
    # RdYlGn: dz<0 (corte) → rojo, dz>0 (relleno) → verde
    cmap = plt.cm.RdYlGn
    im = ax.imshow(delta, cmap=cmap, vmin=-vmax, vmax=vmax,
                   extent=extent, origin="upper", aspect="equal")

    cbar = plt.colorbar(im, ax=ax, shrink=0.7, pad=0.02)
    cbar.set_label("ΔZ  Post − Pre  (m)", fontsize=11)
    ax.set_title("Mapa de Calor — Diferencia de Elevación\n(verde = relleno | rojo = corte)",
                 fontsize=13, fontweight="bold")
    ax.set_xlabel("Este (m)"); ax.set_ylabel("Norte (m)")
    plt.tight_layout()
    plt.savefig(ruta_salida, dpi=150, bbox_inches="tight")
    plt.close()
    print(f"Heatmap guardado: {ruta_salida}")


def extraer_secciones(delta, perfil, eje_path, espaciado, ancho, prog_inicio):
    """Extrae secciones transversales perpendiculares al eje y calcula volúmenes."""
    import geopandas as gpd
    from shapely.geometry import LineString, box
    from shapely.affinity import rotate, translate
    import rasterio
    from rasterio.mask import mask as rmask
    import json

    eje_gdf = gpd.read_file(eje_path)
    if eje_gdf.crs and str(eje_gdf.crs) != str(rasterio.open.__doc__):
        pass  # se puede reproyectar si es necesario

    linea = eje_gdf.geometry.iloc[0]
    if linea.geom_type == "MultiLineString":
        from shapely.ops import linemerge
        linea = linemerge(linea)

    longitud_total = linea.length
    resultados = []

    import rasterio as rio
    with rio.open(os.devnull if not os.path.exists("_tmp_delta.tif") else "_tmp_delta.tif") as _:
        pass

    # Guardar delta temporalmente para usar rasterio.mask
    import tempfile, rasterio
    tmp = tempfile.NamedTemporaryFile(suffix=".tif", delete=False)
    tmp.close()
    perfil_tmp = perfil.copy()
    perfil_tmp.update(dtype="float32", nodata=np.nan)
    with rasterio.open(tmp.name, "w", **perfil_tmp) as dst:
        dst.write(delta.astype(np.float32), 1)

    pixel_area = abs(perfil["transform"].a * perfil["transform"].e)

    distancias = np.arange(0, longitud_total, espaciado)
    with rasterio.open(tmp.name) as src_delta:
        for dist in distancias:
            pt = linea.interpolate(dist)
            pt_next = linea.interpolate(min(dist + 1.0, longitud_total))
            dx = pt_next.x - pt.x
            dy = pt_next.y - pt.y
            ang = np.degrees(np.arctan2(dy, dx))

            # Rectángulo perpendicular al eje
            rect = box(pt.x - espaciado/2, pt.y - ancho, pt.x + espaciado/2, pt.y + ancho)
            rect = rotate(rect, -ang, origin=(pt.x, pt.y))

            try:
                vals, _ = rmask(src_delta, [rect.__geo_interface__], crop=True)
                vals = vals[0]
                vals = vals[~np.isnan(vals)]
                if len(vals) == 0:
                    continue
                corte   = float(np.sum(vals[vals < 0] * -1) * pixel_area)
                relleno = float(np.sum(vals[vals > 0]) * pixel_area)
            except Exception:
                corte, relleno = 0.0, 0.0

            prog = prog_inicio + dist
            resultados.append({
                "progresiva":    round(prog, 2),
                "vol_corte_m3":  round(corte, 3),
                "vol_relleno_m3": round(relleno, 3),
                "vol_esponj_m3": round(corte * FACT_ESPONJ, 3),
                "vol_compac_m3": round(relleno / FACT_CONTRAC, 3),
            })

    os.unlink(tmp.name)
    return resultados


def generar_reporte_excel(resultados, ruta):
    """Genera reporte Excel con tabla por progresiva y resumen."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.chart import LineChart, Reference
    from datetime import date

    wb  = openpyxl.Workbook()
    ws  = wb.active
    ws.title = "Cuantificación ODM"
    ws.sheet_view.showGridLines = False

    hf = PatternFill("solid", fgColor="1F4E79")
    ff = Font(bold=True, color="FFFFFF", size=10)
    thin   = Side(style="thin", color="BFBFBF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center")

    headers = ["Progresiva (m)", "Corte banco\n(m³)", "Relleno\n(m³)",
               "Vol. Esponj.\n(m³)", "Vol. Compac.\n(m³)",
               "Corte Acum.\n(m³)", "Relleno Acum.\n(m³)", "Curva Masa\n(m³)"]
    anchos  = [14, 15, 13, 15, 15, 15, 16, 14]

    ws.row_dimensions[1].height = 35
    for j, (h, w) in enumerate(zip(headers, anchos), start=1):
        ws.column_dimensions[get_column_letter(j)].width = w
        c = ws.cell(row=1, column=j, value=h)
        c.fill = hf; c.font = ff; c.alignment = Alignment(wrap_text=True, horizontal="center", vertical="center")
        c.border = border

    alt = PatternFill("solid", fgColor="EBF3FB")
    acum_corte = 0; acum_relleno = 0; curva_masa = 0

    for i, r in enumerate(resultados):
        f = i + 2
        fill = alt if i % 2 == 0 else None
        acum_corte   += r["vol_corte_m3"]
        acum_relleno += r["vol_relleno_m3"]
        curva_masa   += r["vol_corte_m3"] - r["vol_relleno_m3"]

        vals = [r["progresiva"], r["vol_corte_m3"], r["vol_relleno_m3"],
                r["vol_esponj_m3"], r["vol_compac_m3"],
                round(acum_corte, 3), round(acum_relleno, 3), round(curva_masa, 3)]
        for j, v in enumerate(vals, start=1):
            c = ws.cell(row=f, column=j, value=v)
            c.border = border
            c.number_format = '#,##0.000'
            if fill: c.fill = fill

    # Gráfica curva de masa
    n = len(resultados)
    lc = LineChart()
    lc.title = "Curva de Masa (Flujo Dron+ODM)"
    lc.style = 10
    lc.y_axis.title = "Volumen acumulado (m³)"
    lc.x_axis.title = "Progresiva (m)"
    lc.width = 22; lc.height = 13
    data = Reference(ws, min_col=8, min_row=1, max_row=n+1)
    cats = Reference(ws, min_col=1, min_row=2, max_row=n+1)
    lc.add_data(data, titles_from_data=True)
    lc.set_categories(cats)
    ws.add_chart(lc, "J2")

    # Hoja resumen
    ws2 = wb.create_sheet("Resumen")
    ws2.sheet_view.showGridLines = False
    total_corte   = sum(r["vol_corte_m3"]   for r in resultados)
    total_relleno = sum(r["vol_relleno_m3"] for r in resultados)
    filas = [
        ("Fecha de proceso", str(date.today())),
        ("Total secciones",  len(resultados)),
        ("Espaciado secciones (m)", ESPACIADO_M),
        ("Ancho corredor (m)", ANCHO_CORREDOR_M),
        ("Volumen de corte banco (m³)", round(total_corte, 3)),
        ("Volumen de relleno (m³)", round(total_relleno, 3)),
        ("Volumen esponjado (m³)", round(total_corte * FACT_ESPONJ, 3)),
        ("Balance neto (m³)", round(total_corte - total_relleno, 3)),
        ("Factor esponjamiento", FACT_ESPONJ),
        ("Factor contracción", FACT_CONTRAC),
    ]
    for i, (k, v) in enumerate(filas, start=2):
        ws2.cell(row=i, column=2, value=k).font = Font(bold=True, color="1F4E79")
        ws2.cell(row=i, column=3, value=v)

    wb.save(ruta)
    print(f"Reporte Excel guardado: {ruta}")
    print(f"  Corte total  : {total_corte:.2f} m³")
    print(f"  Relleno total: {total_relleno:.2f} m³")


def exportar_landxml(datos_pre, datos_post, perfil, ruta_salida):
    """
    Exporta las superficies Pre_Obra y Post_Obra como TIN en formato LandXML.
    Importar en Civil 3D: Insertar → Importar LandXML
    """
    import rasterio.transform as rt

    transform = perfil["transform"]
    px = abs(transform.a)   # tamaño de píxel en X
    py = abs(transform.e)   # tamaño de píxel en Y
    paso = PASO_MUESTREO

    def raster_a_puntos(datos, nodata):
        puntos = []
        pid = 1
        for row in range(0, datos.shape[0], paso):
            for col in range(0, datos.shape[1], paso):
                z = float(datos[row, col])
                if nodata is not None and abs(z - nodata) < 1:
                    continue
                if np.isnan(z):
                    continue
                x, y = rt.xy(transform, row, col)
                puntos.append((pid, round(x, 3), round(y, 3), round(z, 4)))
                pid += 1
        return puntos

    def puntos_a_triangulos(puntos):
        """Triangulación simple por cuadrícula."""
        n_cols = datos_pre.shape[1] // paso
        triangulos = []
        n = len(puntos)
        for i in range(n - n_cols - 1):
            if (i + 1) % n_cols == 0:
                continue
            a, b, c, d = i+1, i+2, i+n_cols+1, i+n_cols+2
            if max(a,b,c,d) <= n:
                triangulos.append((a, b, c))
                triangulos.append((b, d, c))
        return triangulos

    nodata = perfil.get("nodata")
    pts_pre  = raster_a_puntos(datos_pre, nodata)
    pts_post = raster_a_puntos(datos_post, nodata)
    tris_pre  = puntos_a_triangulos(pts_pre)
    tris_post = puntos_a_triangulos(pts_post)

    def superficie_xml(nombre, puntos, triangulos):
        pts_xml = "\n".join(
            f'        <P id="{p[0]}">{p[1]} {p[2]} {p[3]}</P>'
            for p in puntos
        )
        tris_xml = "\n".join(
            f'        <F>{t[0]} {t[1]} {t[2]}</F>'
            for t in triangulos
        )
        return f"""  <Surfaces>
    <Surface name="{nombre}">
      <Definition surfType="TIN">
        <Pnts>
{pts_xml}
        </Pnts>
        <Faces>
{tris_xml}
        </Faces>
      </Definition>
    </Surface>
  </Surfaces>"""

    xml = '<?xml version="1.0" encoding="UTF-8"?>\n'
    xml += '<LandXML version="1.2" xmlns="http://www.landxml.org/schema/LandXML-1.2">\n'
    xml += superficie_xml("Pre_Obra", pts_pre, tris_pre) + "\n"
    xml += superficie_xml("Post_Obra", pts_post, tris_post) + "\n"
    xml += "</LandXML>\n"

    with open(ruta_salida, "w", encoding="utf-8") as f:
        f.write(xml)

    size_mb = os.path.getsize(ruta_salida) / 1e6
    print(f"LandXML guardado: {ruta_salida}  ({size_mb:.1f} MB)")
    print(f"  Puntos Pre  : {len(pts_pre):,}")
    print(f"  Puntos Post : {len(pts_post):,}")
    print("  Importar en Civil 3D: Insertar → Importar LandXML")
    print("  Superficies creadas : Pre_Obra  |  Post_Obra")


def main():
    verificar_dependencias()
    script_dir = os.path.dirname(os.path.abspath(__file__))

    dem_pre_path  = os.path.join(script_dir, DEM_PRE)
    dem_post_path = os.path.join(script_dir, DEM_POST)
    eje_path_raw  = os.path.join(script_dir, EJE_VIA)

    for f in [dem_pre_path, dem_post_path, eje_path_raw]:
        if not os.path.exists(f):
            print(f"Archivo no encontrado: {f}")
            sys.exit(1)

    print("=" * 55)
    print(" Cálculo de Volúmenes — Dron + OpenDroneMap")
    print(f" Proyección: EPSG:{EPSG_PROYECTO}")
    print("=" * 55)

    print("[0/5] Cargando eje de vía...")
    eje_path = cargar_eje(eje_path_raw, EPSG_PROYECTO)

    print("[1/5] Alineando DEMs...")
    datos_pre, datos_post, perfil, nodata = alinear_dems(dem_pre_path, dem_post_path)

    print("[2/5] Calculando raster ΔZ...")
    ruta_difdem = os.path.join(script_dir, ARCHIVO_DIFDEM)
    delta = calcular_diferencia(datos_pre, datos_post, perfil, nodata, ruta_difdem)

    print("[3/5] Generando heatmap...")
    ruta_hm = os.path.join(script_dir, ARCHIVO_HEATMAP)
    generar_heatmap(delta, perfil, ruta_hm)

    print("[4/5] Extrayendo secciones transversales...")
    resultados = extraer_secciones(delta, perfil, eje_path,
                                   ESPACIADO_M, ANCHO_CORREDOR_M, PROG_INICIO_M)

    ruta_reporte = os.path.join(script_dir, ARCHIVO_REPORTE)
    generar_reporte_excel(resultados, ruta_reporte)

    print("[5/5] Exportando LandXML para Civil 3D...")
    ruta_xml = os.path.join(script_dir, ARCHIVO_LANDXML)
    exportar_landxml(datos_pre, datos_post, perfil, ruta_xml)

    print("\n✓ Proceso completado.")
    print(f"  {ARCHIVO_REPORTE}")
    print(f"  {ARCHIVO_DIFDEM}")
    print(f"  {ARCHIVO_HEATMAP}")
    print(f"  {ARCHIVO_LANDXML}")


if __name__ == "__main__":
    main()
