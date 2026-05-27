"""
Módulo de exportación de reportes a Excel.

Genera un archivo .xlsx multi-pestaña:
    - Resumen Diario: KPIs del día (volúmenes, balance, áreas).
    - Rendimientos:   Tabla detallada por equipo y resumen por tipo de máquina.
    - Pavimentos:     Cubicación de capas estructurales (si aplica).
    - Histórico:      Registro acumulado de todos los días (apto para dinámicas).

Además, persiste el histórico en disco para acumular jornadas.
"""
from __future__ import annotations

from datetime import date
from io import BytesIO
from pathlib import Path
from typing import Optional

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
TITLE_FONT = Font(bold=True, size=14, color="1F4E78")


def _format_sheet(worksheet, header_row: int = 1) -> None:
    """Aplica formato corporativo a una hoja (encabezado, anchos, alineación)."""
    for cell in worksheet[header_row]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for column_cells in worksheet.columns:
        length = max(
            (len(str(cell.value)) for cell in column_cells if cell.value is not None),
            default=10,
        )
        worksheet.column_dimensions[
            get_column_letter(column_cells[0].column)
        ].width = min(length + 4, 35)


def _load_history(history_path: Path) -> pd.DataFrame:
    """Carga el histórico acumulado si existe; si no, retorna DataFrame vacío."""
    if history_path.exists():
        try:
            return pd.read_excel(history_path)
        except Exception:
            return pd.DataFrame()
    return pd.DataFrame()


def _save_history(history_df: pd.DataFrame, history_path: Path) -> None:
    """Persiste el histórico a disco para conservarlo entre sesiones."""
    history_path.parent.mkdir(parents=True, exist_ok=True)
    history_df.to_excel(history_path, index=False)


def build_daily_summary(
    fecha: date,
    crs_epsg: int,
    cut_fill: dict,
    pavement_results: Optional[dict] = None,
) -> pd.DataFrame:
    """Tabla de KPIs principales del día (la cabecera del reporte)."""
    rows = [
        ("Fecha", fecha.isoformat()),
        ("CRS (EPSG)", crs_epsg),
        ("Volumen de corte (m³)", cut_fill["volumen_corte_m3"]),
        ("Volumen de relleno (m³)", cut_fill["volumen_relleno_m3"]),
        ("Balance neto (m³)", cut_fill["volumen_neto_m3"]),
        ("Área de corte (m²)", cut_fill["area_corte_m2"]),
        ("Área de relleno (m²)", cut_fill["area_relleno_m2"]),
        ("Espesor promedio (m)", cut_fill["espesor_promedio_m"]),
    ]
    if pavement_results:
        for layer, data in pavement_results.items():
            rows.append((f"Pavimento - {layer} (m³)", data["volumen_m3"]))
    return pd.DataFrame(rows, columns=["Métrica", "Valor"])


def export_report(
    output_path: Path,
    history_path: Path,
    fecha: date,
    crs_epsg: int,
    cut_fill: dict,
    yields_df: pd.DataFrame,
    fleet_df: pd.DataFrame,
    pavement_results: Optional[dict] = None,
) -> BytesIO:
    """
    Construye el reporte completo y lo devuelve como BytesIO listo para descarga.

    También actualiza el archivo de histórico acumulado en disco.

    Returns:
        BytesIO con el contenido del .xlsx (para usar en st.download_button).
    """
    # 1. Hojas del día
    summary_df = build_daily_summary(fecha, crs_epsg, cut_fill, pavement_results)

    pavement_df = pd.DataFrame()
    if pavement_results:
        pavement_df = (
            pd.DataFrame.from_dict(pavement_results, orient="index")
            .reset_index()
            .rename(columns={"index": "capa"})
        )

    # 2. Actualizar histórico acumulado
    historic = _load_history(history_path)
    daily_record = yields_df.assign(
        fecha=fecha.isoformat(),
        crs_epsg=crs_epsg,
        volumen_neto_dia_m3=cut_fill["volumen_neto_m3"],
        volumen_corte_dia_m3=cut_fill["volumen_corte_m3"],
        volumen_relleno_dia_m3=cut_fill["volumen_relleno_m3"],
    )
    historic = pd.concat([historic, daily_record], ignore_index=True)
    _save_history(historic, history_path)

    # 3. Escribir libro Excel
    output_path.parent.mkdir(parents=True, exist_ok=True)
    buffer = BytesIO()

    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        summary_df.to_excel(writer, sheet_name="Resumen Diario", index=False)
        yields_df.to_excel(writer, sheet_name="Rendimientos", index=False)
        if not fleet_df.empty:
            fleet_df.to_excel(writer, sheet_name="Resumen Flota", index=False)
        if not pavement_df.empty:
            pavement_df.to_excel(writer, sheet_name="Pavimentos", index=False)
        historic.to_excel(writer, sheet_name="Histórico Acumulado", index=False)

        # Formato a todas las hojas
        for sheet_name in writer.sheets:
            _format_sheet(writer.sheets[sheet_name])

    # Persistir también una copia en reports/ para trazabilidad
    buffer.seek(0)
    with open(output_path, "wb") as f:
        f.write(buffer.getvalue())
    buffer.seek(0)

    return buffer
