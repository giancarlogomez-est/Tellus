"""
Pipeline de seguimiento de avance de obra vial — Dron + ODM
============================================================
Procesa el DSM del día, actualiza el registro histórico y
genera reportes diarios, semanales y mensuales.

Estructura esperada:
    vuelos/YYYY-MM-DD/dsm.tif   ← DSM del día (output de ODM)

Uso:
    python pipeline.py                   # procesa el vuelo de hoy
    python pipeline.py --fecha 2025-05-06
    python pipeline.py --semanal         # fuerza reporte semanal
    python pipeline.py --mensual         # fuerza reporte mensual
============================================================
"""

import argparse
import json
import sys
import warnings
from io import BytesIO
from pathlib import Path
from datetime import date, timedelta

import numpy as np
import pandas as pd
import geopandas as gpd
import ezdxf
import rasterio
from rasterio.warp import reproject, Resampling
from rasterio.mask import mask as rio_mask
from shapely.geometry import LineString, Polygon, mapping
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import matplotlib.colors as mcolors
from matplotlib.patches import Patch
import matplotlib.ticker as mticker
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage

warnings.filterwarnings("ignore")
BASE = Path(__file__).parent

# ══════════════════════════════════════════════════════════════════════════════
# HELPERS EXCEL
# ══════════════════════════════════════════════════════════════════════════════
_L  = Side(style="thin", color="CCCCCC")
BRD = Border(left=_L, right=_L, top=_L, bottom=_L)

def fl(c):   return PatternFill("solid", fgColor=c)
def fn(b=False, c="000000", s=9): return Font(bold=b, color=c, name="Arial", size=s)
def al(h="center", v="center", w=False): return Alignment(horizontal=h, vertical=v, wrap_text=w)

def title_row(ws, r, nc, text, bg="1F2937", sz=11):
    ws.merge_cells(f"A{r}:{get_column_letter(nc)}{r}")
    c = ws.cell(row=r, column=1, value=text)
    c.fill = fl(bg); c.font = fn(True, "FFFFFF", sz)
    c.alignment = al(); ws.row_dimensions[r].height = 26

def hdr(ws, r, nc, bg="374151", h=28):
    for col in range(1, nc+1):
        c = ws.cell(row=r, column=col)
        c.fill = fl(bg); c.font = fn(True, "FFFFFF", 9)
        c.alignment = al(w=True); c.border = BRD
    ws.row_dimensions[r].height = h

def cell(ws, r, c, v, bold=False, bg=None, color="000000", fmt=None, h="center"):
    x = ws.cell(row=r, column=c, value=v)
    x.font = fn(bold, color, 9); x.border = BRD; x.alignment = al(h=h)
    if bg:  x.fill = fl(bg)
    if fmt: x.number_format = fmt
    return x

def metric_block(ws, r0, label, valor, unidad="", bg_label="374151", fmt=None):
    """Bloque de 2 celdas: etiqueta izquierda + valor derecha."""
    cl = ws.cell(row=r0, column=1, value=label)
    cl.fill = fl(bg_label); cl.font = fn(True, "FFFFFF", 9)
    cl.alignment = al(h="left"); cl.border = BRD
    cv = ws.cell(row=r0, column=2, value=valor)
    cv.font = fn(True, "000000", 10); cv.alignment = al(h="right")
    cv.border = BRD
    if fmt: cv.number_format = fmt
    ws.cell(row=r0, column=3, value=unidad).font = fn(False, "666666", 9)
    ws.cell(row=r0, column=3).border = BRD
    ws.cell(row=r0, column=3).alignment = al(h="left")
    ws.row_dimensions[r0].height = 18


# ══════════════════════════════════════════════════════════════════════════════
# CONFIGURACIÓN
# ══════════════════════════════════════════════════════════════════════════════
def cargar_config():
    p = BASE / "proyecto_config.json"
    if not p.exists():
        sys.exit("[ERROR] No se encontró proyecto_config.json — corre configurar_proyecto.py primero.")
    c = json.loads(p.read_text(encoding="utf-8"))
    c["_base"] = BASE
    return c

def ruta(cfg, key):
    return BASE / cfg[key]


# ══════════════════════════════════════════════════════════════════════════════
# REGISTRO CSV
# ══════════════════════════════════════════════════════════════════════════════
REG_COLS = ["fecha","vuelo_num","vol_corte_dia","vol_relleno_dia","balance_dia",
            "vol_corte_acum","vol_relleno_acum","balance_acum","semana","mes"]

SEC_COLS = ["fecha","progresiva","vol_corte_dia","vol_relleno_dia",
            "vol_corte_acum","vol_relleno_acum"]

def cargar_registro(base):
    p = base / "registro.csv"
    if p.exists():
        return pd.read_csv(p, parse_dates=["fecha"])
    return pd.DataFrame(columns=REG_COLS)

def cargar_registro_secciones(base):
    p = base / "registro_secciones.csv"
    if p.exists():
        return pd.read_csv(p, parse_dates=["fecha"])
    return pd.DataFrame(columns=SEC_COLS)

def guardar_registro(base, df):
    df.to_csv(base / "registro.csv", index=False, date_format="%Y-%m-%d")

def guardar_registro_secciones(base, df):
    df.to_csv(base / "registro_secciones.csv", index=False, date_format="%Y-%m-%d")


# ══════════════════════════════════════════════════════════════════════════════
# FUNCIONES DE CÁLCULO RASTER
# ══════════════════════════════════════════════════════════════════════════════
def cargar_dem(path, nodata=-9999.0):
    with rasterio.open(path) as src:
        arr = src.read(1).astype("float64")
        arr[arr == (src.nodata if src.nodata else nodata)] = np.nan
        return arr, src.transform, src.crs, src.meta.copy()

def alinear_dem(path, ref_tf, ref_crs, ref_shape, nodata=-9999.0):
    arr = np.full(ref_shape, np.nan, dtype="float64")
    with rasterio.open(path) as src:
        reproject(source=rasterio.band(src, 1), destination=arr,
                  src_transform=src.transform, src_crs=src.crs,
                  dst_transform=ref_tf, dst_crs=ref_crs,
                  resampling=Resampling.bilinear,
                  src_nodata=src.nodata or nodata, dst_nodata=np.nan)
    return arr

def cargar_eje_dxf(path, crs_target=None):
    """
    Carga el eje de la vía desde un archivo DXF.

    Entidades soportadas: LINE, LWPOLYLINE, POLYLINE, SPLINE, ARC.
    Si el DXF tiene varias entidades, las une en una sola LineString
    ordenándolas por conectividad (extremo más cercano).

    Parámetros
    ----------
    path       : str | Path — ruta al archivo .dxf
    crs_target : rasterio.crs.CRS — CRS del raster de referencia.
                 Si el DXF tiene $INSUNITS en metros se asume que
                 sus coordenadas ya están en ese CRS.
                 Si no coincide, reproyecta con pyproj si el DXF
                 tiene CRS definido, o lanza advertencia.

    Retorna
    -------
    shapely.geometry.LineString
    """
    from shapely.geometry import LineString, MultiLineString
    from shapely.ops import linemerge, unary_union
    import ezdxf

    doc = ezdxf.readfile(str(path))
    msp = doc.modelspace()

    segmentos = []

    for ent in msp:
        tipo = ent.dxftype()

        if tipo == "LINE":
            p1 = (ent.dxf.start.x, ent.dxf.start.y)
            p2 = (ent.dxf.end.x,   ent.dxf.end.y)
            segmentos.append(LineString([p1, p2]))

        elif tipo == "LWPOLYLINE":
            pts = [(p[0], p[1]) for p in ent.get_points()]
            if len(pts) >= 2:
                segmentos.append(LineString(pts))

        elif tipo == "POLYLINE":
            pts = [(v.dxf.location.x, v.dxf.location.y)
                   for v in ent.vertices]
            if len(pts) >= 2:
                segmentos.append(LineString(pts))

        elif tipo == "SPLINE":
            # Aproximar la spline con puntos muestreados
            try:
                pts = [(p.x, p.y) for p in ent.flattening(0.01)]
            except Exception:
                pts = [(p[0], p[1]) for p in ent.control_points]
            if len(pts) >= 2:
                segmentos.append(LineString(pts))

        elif tipo == "ARC":
            import math
            cx, cy = ent.dxf.center.x, ent.dxf.center.y
            r      = ent.dxf.radius
            a_ini  = math.radians(ent.dxf.start_angle)
            a_fin  = math.radians(ent.dxf.end_angle)
            if a_fin < a_ini:
                a_fin += 2 * math.pi
            n_pts = max(32, int((a_fin - a_ini) / math.radians(2)))
            angles = np.linspace(a_ini, a_fin, n_pts)
            pts = [(cx + r*math.cos(a), cy + r*math.sin(a)) for a in angles]
            segmentos.append(LineString(pts))

    if not segmentos:
        raise ValueError(
            f"No se encontraron entidades LINE/LWPOLYLINE/POLYLINE/SPLINE/ARC "
            f"en '{path}'. Verifica que el eje esté en el modelspace."
        )

    # Unir todos los segmentos en una sola línea
    if len(segmentos) == 1:
        linea = segmentos[0]
    else:
        merged = linemerge(unary_union(segmentos))
        if merged.geom_type == "MultiLineString":
            # Si no se pudo unir completamente, usar el segmento más largo
            linea = max(merged.geoms, key=lambda g: g.length)
            print(f"  [!] DXF contiene {len(segmentos)} segmentos discontinuos — "
                  f"se usó el más largo ({linea.length:.1f} m). "
                  f"Verifica que el eje sea una polilínea continua.")
        else:
            linea = merged

    print(f"  Eje DXF cargado: {linea.length:.1f} m  "
          f"({len(segmentos)} entidad(es) leída(s))")
    return linea
    secs, hw = [], ancho
    dists = np.arange(0, linea.length, espaciado)
    for d in dists:
        d2 = min(d + espaciado, linea.length)
        p1 = linea.interpolate(d); p2 = linea.interpolate(d2)
        dx, dy = p2.x-p1.x, p2.y-p1.y
        mag = np.hypot(dx, dy)
        if mag < 1e-6: continue
        nx, ny = -dy/mag, dx/mag
        poly = Polygon([(p1.x+nx*hw, p1.y+ny*hw), (p2.x+nx*hw, p2.y+ny*hw),
                        (p2.x-nx*hw, p2.y-ny*hw), (p1.x-nx*hw, p1.y-ny*hw)])
        dm = prog_ini + d
        secs.append({"progresiva": f"K{int(dm//1000)}+{int(dm%1000):03d}",
                     "dist": round(d, 1), "polygon": poly})
    return secs

def crear_secciones(linea, espaciado, ancho, prog_ini):
    secs, hw = [], ancho
    dists = np.arange(0, linea.length, espaciado)
    for d in dists:
        d2 = min(d + espaciado, linea.length)
        p1 = linea.interpolate(d); p2 = linea.interpolate(d2)
        dx, dy = p2.x-p1.x, p2.y-p1.y
        mag = np.hypot(dx, dy)
        if mag < 1e-6: continue
        nx, ny = -dy/mag, dx/mag
        poly = Polygon([(p1.x+nx*hw, p1.y+ny*hw), (p2.x+nx*hw, p2.y+ny*hw),
                        (p2.x-nx*hw, p2.y-ny*hw), (p1.x-nx*hw, p1.y-ny*hw)])
        dm = prog_ini + d
        secs.append({"progresiva": f"K{int(dm//1000)}+{int(dm%1000):03d}",
                     "dist": round(d, 1), "polygon": poly})
    return secs

def calcular_volumenes(dz_arr, tf, secciones, pixel_area, dz_raster_path):
    rows = []
    with rasterio.open(dz_raster_path) as src:
        for s in secciones:
            try:
                m, _ = rio_mask(src, [mapping(s["polygon"])],
                                crop=True, nodata=-9999.0, all_touched=True)
                a = m[0].astype("float64"); a[a == -9999.0] = np.nan
                v = a[np.isfinite(a)]
                vc = float(abs(v[v < 0].sum()) * pixel_area)
                vr = float(v[v >= 0].sum() * pixel_area)
            except Exception:
                vc = vr = 0.0
            rows.append({"progresiva": s["progresiva"], "vol_corte": round(vc, 3),
                         "vol_relleno": round(vr, 3)})
    return pd.DataFrame(rows)

def guardar_dz_raster(dz_arr, meta, path):
    m = meta.copy(); m.update(dtype="float32", nodata=-9999.0)
    arr = np.where(np.isfinite(dz_arr), dz_arr, -9999.0).astype("float32")
    with rasterio.open(path, "w", **m) as dst:
        dst.write(arr, 1)


# ══════════════════════════════════════════════════════════════════════════════
# HEATMAP
# ══════════════════════════════════════════════════════════════════════════════
def generar_heatmap(dz_arr, tf, secciones, eje, titulo, ruta_png,
                    dz_min=-3.0, dz_max=3.0):
    with rasterio.open(ruta_png.replace(".png", "_tmp.tif"),
                       "w", driver="GTiff", dtype="float32", nodata=-9999.0,
                       width=dz_arr.shape[1], height=dz_arr.shape[0],
                       count=1, crs=None, transform=tf) as _:
        pass  # solo para obtener bounds
    bounds = rasterio.transform.array_bounds(
        dz_arr.shape[0], dz_arr.shape[1], tf)
    extent = [bounds[0], bounds[2], bounds[1], bounds[3]]

    fig, ax = plt.subplots(figsize=(13, 4.5))
    cmap = plt.get_cmap("RdBu")
    norm = mcolors.TwoSlopeNorm(vmin=dz_min, vcenter=0, vmax=dz_max)
    ax.imshow(dz_arr, extent=extent, cmap=cmap, norm=norm,
              aspect="equal", origin="upper", interpolation="bilinear")

    for s in secciones:
        xp, yp = s["polygon"].exterior.xy
        ax.plot(xp, yp, color="gray", lw=0.5, alpha=0.5)

    ex, ey = eje.xy
    ax.plot(ex, ey, "k-", lw=1.6, label="Eje")

    cbar = plt.colorbar(
        plt.cm.ScalarMappable(norm=norm, cmap=cmap),
        ax=ax, fraction=0.025, pad=0.02)
    cbar.set_label("ΔZ (m)", fontsize=8)
    legend_el = [Patch(facecolor="#3B82F6", label="Relleno"),
                 Patch(facecolor="#EF4444", label="Corte")]
    ax.legend(handles=legend_el+[plt.Line2D([0],[0],color="k",lw=1.6,label="Eje")],
              fontsize=7, loc="upper left")
    ax.set_title(titulo, fontsize=9)
    ax.tick_params(labelsize=7)
    fig.tight_layout()
    fig.savefig(ruta_png, dpi=150, bbox_inches="tight")
    plt.close(fig)
    Path(ruta_png.replace(".png", "_tmp.tif")).unlink(missing_ok=True)


# ══════════════════════════════════════════════════════════════════════════════
# REPORTE DIARIO
# ══════════════════════════════════════════════════════════════════════════════
def reporte_diario(cfg, fecha, vuelo_num, df_sec_dia, df_sec_acum,
                   reg, heatmap_path, ruta_xlsx):
    wb = Workbook()
    nom = cfg["nombre"]; F2 = "#,##0.00"

    tot_c_dia  = df_sec_dia["vol_corte"].sum()
    tot_r_dia  = df_sec_dia["vol_relleno"].sum()
    bal_dia    = tot_c_dia - tot_r_dia
    tot_c_acum = df_sec_acum["vol_corte_acum"].sum()
    tot_r_acum = df_sec_acum["vol_relleno_acum"].sum()
    bal_acum   = tot_c_acum - tot_r_acum
    pct_c = (tot_c_acum / cfg["vol_corte_objetivo"] * 100) if cfg["vol_corte_objetivo"] else 0
    pct_r = (tot_r_acum / cfg["vol_relleno_objetivo"] * 100) if cfg["vol_relleno_objetivo"] else 0

    # ── Hoja 1: Resumen ────────────────────────────────────────────────────
    ws = wb.active; ws.title = "Resumen"
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 10

    title_row(ws, 1, 3, f"Informe Diario de Avance — {fecha.strftime('%d/%m/%Y')}")
    ws.cell(row=2, column=1, value=f"Proyecto: {nom}").font = fn(False, "444444", 9)
    ws.cell(row=2, column=2, value=f"Vuelo N° {vuelo_num}").font = fn(True, "000000", 9)
    ws.cell(row=2, column=2).alignment = al(h="right")
    ws.cell(row=2, column=3, value=fecha.strftime("Sem. %W/%Y")).font = fn(False, "666666", 9)
    ws.row_dimensions[2].height = 16

    ws.cell(row=4, column=1, value="PRODUCCIÓN DEL DÍA").fill = fl("1F2937")
    ws.cell(row=4, column=1).font = fn(True, "FFFFFF", 9)
    ws.cell(row=4, column=1).alignment = al(h="left")
    for c in [2, 3]: ws.cell(row=4, column=c).fill = fl("1F2937")
    ws.row_dimensions[4].height = 18

    metric_block(ws, 5,  "Volumen de Corte",   round(tot_c_dia, 2),  "m³", "3B4E6E")
    metric_block(ws, 6,  "Volumen de Relleno", round(tot_r_dia, 2),  "m³", "6E3B3B")
    metric_block(ws, 7,  "Balance (C − R)",    round(bal_dia, 2),    "m³",
                 "1D4ED8" if bal_dia > 0 else "B91C1C")

    ws.cell(row=9, column=1, value="ACUMULADO HASTA LA FECHA").fill = fl("1F2937")
    ws.cell(row=9, column=1).font = fn(True, "FFFFFF", 9)
    ws.cell(row=9, column=1).alignment = al(h="left")
    for c in [2, 3]: ws.cell(row=9, column=c).fill = fl("1F2937")
    ws.row_dimensions[9].height = 18

    metric_block(ws, 10, "Corte acumulado",    round(tot_c_acum, 2), "m³", "3B4E6E")
    metric_block(ws, 11, "Relleno acumulado",  round(tot_r_acum, 2), "m³", "6E3B3B")
    metric_block(ws, 12, "Balance acumulado",  round(bal_acum, 2),   "m³",
                 "1D4ED8" if bal_acum > 0 else "B91C1C")
    metric_block(ws, 13, "% Avance corte",     round(pct_c, 1), "% del objetivo", "2D6A4F")
    metric_block(ws, 14, "% Avance relleno",   round(pct_r, 1), "% del objetivo", "2D6A4F")

    # Heatmap
    if heatmap_path and Path(heatmap_path).exists():
        img = XLImage(heatmap_path)
        img.width = 620; img.height = 215
        ws.add_image(img, "E2")

    # ── Hoja 2: Secciones ─────────────────────────────────────────────────
    ws2 = wb.create_sheet("Secciones")
    ws2.sheet_view.showGridLines = False
    cols = ["Progresiva", "Corte hoy (m³)", "Relleno hoy (m³)", "Balance hoy (m³)",
            "Corte acum. (m³)", "Relleno acum. (m³)", "Balance acum. (m³)"]
    anchos = [12, 15, 16, 15, 15, 16, 16]
    for i, w in enumerate(anchos, 1):
        ws2.column_dimensions[get_column_letter(i)].width = w

    title_row(ws2, 1, 7, f"Cuantificación por Sección — {fecha.strftime('%d/%m/%Y')}")
    for c, h in enumerate(cols, 1): ws2.cell(row=2, column=c).value = h
    hdr(ws2, 2, 7)

    df_merged = df_sec_dia.merge(df_sec_acum, on="progresiva", how="left").fillna(0)
    for i, row in df_merged.iterrows():
        r = i + 3
        b_dia  = row["vol_corte"] - row["vol_relleno"]
        b_acum = row.get("vol_corte_acum", 0) - row.get("vol_relleno_acum", 0)
        bg = fl("F9FAFB") if i % 2 == 0 else fl("FFFFFF")
        vals = [row["progresiva"], row["vol_corte"], row["vol_relleno"], b_dia,
                row.get("vol_corte_acum", 0), row.get("vol_relleno_acum", 0), b_acum]
        for c, v in enumerate(vals, 1):
            x = ws2.cell(row=r, column=c, value=v)
            x.font = fn(s=9); x.border = BRD; x.alignment = al(); x.fill = bg
            if c > 1: x.number_format = F2

    # Fila totales
    ft = len(df_merged) + 3
    n  = len(df_merged)
    for c in range(1, 8):
        x = ws2.cell(row=ft, column=c)
        x.fill = fl("1F2937"); x.font = fn(True, "FFFFFF", 9)
        x.border = BRD; x.alignment = al()
    ws2.cell(row=ft, column=1).value = "TOTALES"
    for c, col in [(2,"B"),(3,"C"),(4,"D"),(5,"E"),(6,"F"),(7,"G")]:
        ws2.cell(row=ft, column=c).value = f"=SUM({col}3:{col}{n+2})"
        ws2.cell(row=ft, column=c).number_format = F2

    ws2.freeze_panes = "A3"
    wb.save(ruta_xlsx)


# ══════════════════════════════════════════════════════════════════════════════
# REPORTE SEMANAL
# ══════════════════════════════════════════════════════════════════════════════
def reporte_semanal(cfg, semana_df, semana_label, ruta_xlsx):
    """semana_df: filas del registro para los días de esa semana."""
    wb = Workbook(); F2 = "#,##0.00"
    ws = wb.active; ws.title = "Resumen Semanal"
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 16
    for i in range(2, 7): ws.column_dimensions[get_column_letter(i)].width = 14

    title_row(ws, 1, 6, f"Informe Semanal de Avance — {semana_label}")
    ws.cell(row=2, column=1, value=f"Proyecto: {cfg['nombre']}").font = fn(False, "444444", 9)
    ws.row_dimensions[2].height = 16

    cols = ["Fecha", "Vuelo N°", "Corte día (m³)", "Relleno día (m³)",
            "Balance día (m³)", "Acum. corte (m³)"]
    for c, h in enumerate(cols, 1): ws.cell(row=4, column=c).value = h
    hdr(ws, 4, 6)

    for i, row in semana_df.reset_index(drop=True).iterrows():
        r = i + 5
        bg = fl("F9FAFB") if i % 2 == 0 else fl("FFFFFF")
        vals = [pd.Timestamp(row["fecha"]).strftime("%d/%m/%Y"),
                int(row["vuelo_num"]),
                row["vol_corte_dia"], row["vol_relleno_dia"],
                row["balance_dia"], row["vol_corte_acum"]]
        for c, v in enumerate(vals, 1):
            x = ws.cell(row=r, column=c, value=v)
            x.font = fn(s=9); x.border = BRD; x.alignment = al(); x.fill = bg
            if c > 2: x.number_format = F2

    ft = len(semana_df) + 5
    ws.merge_cells(f"A{ft}:B{ft}")
    for c in range(1, 7):
        x = ws.cell(row=ft, column=c)
        x.fill = fl("1F2937"); x.font = fn(True, "FFFFFF", 9)
        x.border = BRD; x.alignment = al()
    ws.cell(row=ft, column=1).value = "TOTALES"
    nd = len(semana_df)
    for c, col in [(3,"C"),(4,"D"),(5,"E")]:
        ws.cell(row=ft, column=c).value = f"=SUM({col}5:{col}{nd+4})"
        ws.cell(row=ft, column=c).number_format = F2

    # Gráfico de barras diario
    fig, (ax1, ax2) = plt.subplots(2, 1, figsize=(10, 6))
    fechas = [pd.Timestamp(f).strftime("%d/%m") for f in semana_df["fecha"]]
    xi = range(len(fechas))

    ax1.bar(xi,  semana_df["vol_corte_dia"].values,   label="Corte",   color="#3B4E6E", alpha=0.85)
    ax1.bar(xi, -semana_df["vol_relleno_dia"].values,  label="Relleno", color="#6E3B3B", alpha=0.85)
    ax1.axhline(0, color="black", lw=0.8)
    ax1.set_xticks(xi); ax1.set_xticklabels(fechas, fontsize=8)
    ax1.set_title("Producción diaria de la semana", fontsize=10)
    ax1.legend(fontsize=8); ax1.grid(axis="y", ls="--", alpha=0.4)
    ax1.yaxis.set_major_formatter(mticker.FuncFormatter(lambda v,_: f"{abs(v):,.0f}"))

    ax2.plot(xi, semana_df["vol_corte_acum"].values, "b-o", ms=5, lw=1.5, label="Corte acum.")
    ax2.plot(xi, semana_df["vol_relleno_acum"].values, "r-o", ms=5, lw=1.5, label="Relleno acum.")
    ax2.set_xticks(xi); ax2.set_xticklabels(fechas, fontsize=8)
    ax2.set_title("Acumulado al cierre de cada día", fontsize=10)
    ax2.legend(fontsize=8); ax2.grid(axis="y", ls="--", alpha=0.4)
    ax2.yaxis.set_major_formatter(mticker.FuncFormatter(lambda v,_: f"{v:,.0f}"))
    fig.tight_layout()

    buf = BytesIO(); fig.savefig(buf, format="png", dpi=130, bbox_inches="tight"); buf.seek(0)
    plt.close(fig)
    img = XLImage(buf); img.width = 720; img.height = 430
    ws.add_image(img, f"H4")

    wb.save(ruta_xlsx)


# ══════════════════════════════════════════════════════════════════════════════
# REPORTE MENSUAL
# ══════════════════════════════════════════════════════════════════════════════
def reporte_mensual(cfg, mes_df, mes_label, heatmap_path, ruta_xlsx):
    wb = Workbook(); F2 = "#,##0.00"
    ws = wb.active; ws.title = "Resumen Mensual"
    ws.sheet_view.showGridLines = False
    for i, w in enumerate([14,10,15,15,15,15,10], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    title_row(ws, 1, 7, f"Informe Mensual de Avance — {mes_label}")
    ws.cell(row=2, column=1, value=f"Proyecto: {cfg['nombre']}").font = fn(False,"444444",9)

    cols = ["Fecha","Vuelo","Corte día (m³)","Relleno día (m³)","Balance día (m³)",
            "Corte acum. (m³)","% Avance corte"]
    for c, h in enumerate(cols, 1): ws.cell(row=4, column=c).value = h
    hdr(ws, 4, 7)

    for i, row in mes_df.reset_index(drop=True).iterrows():
        r = i + 5
        pct = (row["vol_corte_acum"] / cfg["vol_corte_objetivo"] * 100
               if cfg["vol_corte_objetivo"] else 0)
        bg = fl("F9FAFB") if i%2==0 else fl("FFFFFF")
        vals = [pd.Timestamp(row["fecha"]).strftime("%d/%m/%Y"),
                int(row["vuelo_num"]),
                row["vol_corte_dia"], row["vol_relleno_dia"], row["balance_dia"],
                row["vol_corte_acum"], round(pct, 1)]
        for c, v in enumerate(vals, 1):
            x = ws.cell(row=r, column=c, value=v)
            x.font = fn(s=9); x.border = BRD; x.alignment = al(); x.fill = bg
            if c in [3,4,5,6]: x.number_format = F2
            if c == 7: x.number_format = "0.0%"; x.value = pct/100

    # Curva S
    ws2 = wb.create_sheet("Curva S")
    ws2.sheet_view.showGridLines = False
    title_row(ws2, 1, 4, f"Curva S de Avance — {mes_label}")

    fig, ax = plt.subplots(figsize=(11, 5))
    fechas = [pd.Timestamp(f).strftime("%d/%m") for f in mes_df["fecha"]]
    pcts_c = (mes_df["vol_corte_acum"] / cfg["vol_corte_objetivo"] * 100
              if cfg["vol_corte_objetivo"] else mes_df["vol_corte_acum"] * 0)
    pcts_r = (mes_df["vol_relleno_acum"] / cfg["vol_relleno_objetivo"] * 100
              if cfg["vol_relleno_objetivo"] else mes_df["vol_relleno_acum"] * 0)

    ax.plot(range(len(fechas)), pcts_c.values, "b-o", ms=5, lw=2,
            label=f"Corte (obj: {cfg['vol_corte_objetivo']:,} m³)")
    ax.plot(range(len(fechas)), pcts_r.values, "r-o", ms=5, lw=2,
            label=f"Relleno (obj: {cfg['vol_relleno_objetivo']:,} m³)")
    ax.axhline(100, color="green", lw=1, ls="--", label="100% objetivo")
    ax.set_xticks(range(len(fechas)))
    ax.set_xticklabels(fechas, rotation=45, ha="right", fontsize=7)
    ax.set_ylabel("% de avance vs objetivo"); ax.set_ylim(0, max(110, pcts_c.max()+5))
    ax.set_title(f"Curva S — Avance acumulado | {mes_label}", fontsize=10)
    ax.legend(fontsize=8); ax.grid(ls="--", alpha=0.4)
    fig.tight_layout()
    buf = BytesIO(); fig.savefig(buf, format="png", dpi=130, bbox_inches="tight"); buf.seek(0)
    plt.close(fig)
    img = XLImage(buf); img.width = 780; img.height = 360
    ws2.add_image(img, "F2")

    # Heatmap del mes
    if heatmap_path and Path(heatmap_path).exists():
        ws3 = wb.create_sheet("Mapa de calor mensual")
        ws3.sheet_view.showGridLines = False
        title_row(ws3, 1, 2, f"Mapa de diferencias del mes — {mes_label}")
        img2 = XLImage(heatmap_path); img2.width = 820; img2.height = 290
        ws3.add_image(img2, "A3")

    wb.save(ruta_xlsx)


# ══════════════════════════════════════════════════════════════════════════════
# LANDXML (solo en reporte mensual)
# ══════════════════════════════════════════════════════════════════════════════
def escribir_landxml(superficies, tf, paso, ruta, crs_epsg=None):
    res = abs(tf.a) * paso
    lines = ['<?xml version="1.0" encoding="UTF-8"?>',
             '<LandXML version="1.2" xmlns="http://www.landxml.org/schema/LandXML-1.2"',
             '  xmlns:xsi="http://www.w3.org/2001/XMLSchema-instance">',
             '  <Units><Metric areaUnit="squareMeter" linearUnit="meter" volumeUnit="cubicMeter"/></Units>']
    if crs_epsg:
        lines.append(f'  <!-- EPSG:{crs_epsg} | Resolución malla: {res:.2f} m -->')
    lines.append('  <Surfaces>')

    for nombre, desc, arr in superficies:
        ri = list(range(0, arr.shape[0], paso))
        ci = list(range(0, arr.shape[1], paso))
        id_map = {}; pid = 1; plines = []
        for r in ri:
            for c in ci:
                z = float(arr[r, c])
                if np.isnan(z): continue
                e, n = rasterio.transform.xy(tf, r, c)
                id_map[(r,c)] = pid
                plines.append(f'          <P id="{pid}">{n:.4f} {e:.4f} {z:.4f}</P>')
                pid += 1
        flines = []
        for ri2 in range(len(ri)-1):
            for ci2 in range(len(ci)-1):
                c00,c01 = (ri[ri2],ci[ci2]), (ri[ri2],ci[ci2+1])
                c10,c11 = (ri[ri2+1],ci[ci2]), (ri[ri2+1],ci[ci2+1])
                if all(k in id_map for k in [c00,c01,c10,c11]):
                    p0,p1,p2,p3 = id_map[c00],id_map[c01],id_map[c10],id_map[c11]
                    flines += [f'          <F>{p0} {p1} {p2}</F>',
                               f'          <F>{p1} {p3} {p2}</F>']
        lines += [f'    <Surface name="{nombre}" desc="{desc}">',
                  '      <Definition surfType="TIN"><Pnts>']
        lines += plines
        lines += ['      </Pnts><Faces>']
        lines += flines
        lines += ['      </Faces></Definition></Surface>']
        print(f"    LandXML {nombre}: {pid-1} pts | {len(flines)} caras")

    lines += ['  </Surfaces>', '</LandXML>']
    Path(ruta).write_text("\n".join(lines), encoding="utf-8")


# ══════════════════════════════════════════════════════════════════════════════
# PIPELINE PRINCIPAL
# ══════════════════════════════════════════════════════════════════════════════
def procesar_dia(cfg, fecha):
    base    = cfg["_base"]
    dsm_hoy = base / "vuelos" / fecha.isoformat() / "dsm.tif"
    if not dsm_hoy.exists():
        sys.exit(f"[ERROR] No se encontró el DSM de hoy: {dsm_hoy}")

    baseline_path = ruta(cfg, "dem_baseline")
    eje_path      = ruta(cfg, "eje_via")
    if not baseline_path.exists():
        sys.exit(f"[ERROR] No se encontró baseline: {baseline_path}")

    print(f"\n  Procesando vuelo: {fecha.isoformat()}")

    # Cargar baseline
    arr_base, tf_ref, crs_ref, meta_ref = cargar_dem(str(baseline_path))
    pixel_area = abs(tf_ref.a * tf_ref.e)
    nrows, ncols = arr_base.shape

    # Cargar DSM de hoy (reproyectado a grilla baseline)
    arr_hoy = alinear_dem(str(dsm_hoy), tf_ref, crs_ref, (nrows, ncols))

    # ΔZ vs baseline → acumulado
    dz_acum = arr_hoy - arr_base

    # ΔZ vs ayer → producción del día
    reg = cargar_registro(base)
    reg_sorted = reg.sort_values("fecha")
    dsm_ayer = None
    if not reg_sorted.empty:
        ayer = reg_sorted.iloc[-1]
        dsm_ayer_path = base / "vuelos" / pd.Timestamp(ayer["fecha"]).strftime("%Y-%m-%d") / "dsm.tif"
        if dsm_ayer_path.exists():
            dsm_ayer = alinear_dem(str(dsm_ayer_path), tf_ref, crs_ref, (nrows, ncols))

    dz_dia = (arr_hoy - dsm_ayer) if dsm_ayer is not None else dz_acum

    # Guardar rasters de diferencia
    tmp_acum = str(base / "vuelos" / fecha.isoformat() / "dz_acum.tif")
    tmp_dia  = str(base / "vuelos" / fecha.isoformat() / "dz_dia.tif")
    guardar_dz_raster(dz_acum, meta_ref, tmp_acum)
    guardar_dz_raster(dz_dia,  meta_ref, tmp_dia)

    # Eje y secciones
    gdf = gpd.read_file(str(eje_path)) if str(eje_path).endswith((".geojson", ".shp")) \
          else None
    if gdf is not None:
        if gdf.crs and gdf.crs != crs_ref: gdf = gdf.to_crs(crs_ref)
        eje = gdf.geometry.iloc[0]
    else:
        eje = cargar_eje_dxf(str(eje_path), crs_ref)
    secs = crear_secciones(eje, cfg["espaciado_secciones"],
                           cfg["ancho_corredor"], cfg["prog_inicio_m"])

    # Volúmenes del día y acumulados
    df_dia  = calcular_volumenes(dz_dia,  tf_ref, secs, pixel_area, tmp_dia)
    df_acum = calcular_volumenes(dz_acum, tf_ref, secs, pixel_area, tmp_acum)
    df_acum = df_acum.rename(columns={"vol_corte":"vol_corte_acum",
                                       "vol_relleno":"vol_relleno_acum"})

    tot_c_dia  = df_dia["vol_corte"].sum()
    tot_r_dia  = df_dia["vol_relleno"].sum()
    tot_c_acum = df_acum["vol_corte_acum"].sum()
    tot_r_acum = df_acum["vol_relleno_acum"].sum()
    vuelo_num  = len(reg) + 1

    print(f"  Corte hoy:    {tot_c_dia:>10,.2f} m³  |  Acum: {tot_c_acum:>10,.2f} m³")
    print(f"  Relleno hoy:  {tot_r_dia:>10,.2f} m³  |  Acum: {tot_r_acum:>10,.2f} m³")

    # Actualizar registro
    nueva_fila = pd.DataFrame([{
        "fecha":            fecha,
        "vuelo_num":        vuelo_num,
        "vol_corte_dia":    round(tot_c_dia, 3),
        "vol_relleno_dia":  round(tot_r_dia, 3),
        "balance_dia":      round(tot_c_dia - tot_r_dia, 3),
        "vol_corte_acum":   round(tot_c_acum, 3),
        "vol_relleno_acum": round(tot_r_acum, 3),
        "balance_acum":     round(tot_c_acum - tot_r_acum, 3),
        "semana":           fecha.strftime("%Y-W%W"),
        "mes":              fecha.strftime("%Y-%m"),
    }])
    reg = pd.concat([reg, nueva_fila], ignore_index=True)
    guardar_registro(base, reg)

    # Registro de secciones
    reg_sec = cargar_registro_secciones(base)
    df_sec_new = df_dia.rename(columns={"vol_corte":"vol_corte_dia",
                                         "vol_relleno":"vol_relleno_dia"})
    df_sec_new = df_sec_new.merge(df_acum, on="progresiva")
    df_sec_new["fecha"] = fecha
    df_sec_new = df_sec_new[SEC_COLS]
    reg_sec = pd.concat([reg_sec, df_sec_new], ignore_index=True)
    guardar_registro_secciones(base, reg_sec)

    # Heatmap del día
    dir_d = base / "reportes" / "diarios"
    dir_d.mkdir(parents=True, exist_ok=True)
    heatmap_path = str(dir_d / f"heatmap_{fecha.isoformat()}.png")
    generar_heatmap(dz_dia, tf_ref, secs, eje,
                    f"ΔZ del día {fecha.strftime('%d/%m/%Y')} — producción diaria",
                    heatmap_path,
                    cfg["dz_min_plot"], cfg["dz_max_plot"])

    # Reporte diario
    ruta_xlsx = str(dir_d / f"reporte_{fecha.isoformat()}.xlsx")
    reporte_diario(cfg, fecha, vuelo_num, df_dia, df_acum,
                   reg.iloc[-1], heatmap_path, ruta_xlsx)
    print(f"  → {ruta_xlsx}")
    return reg


def procesar_semanal(cfg, reg, fecha_ref):
    base = cfg["_base"]
    reg = reg.copy()
    reg["fecha"] = pd.to_datetime(reg["fecha"], format='mixed')
    if fecha_ref.weekday() == 0:
        semana_str = (fecha_ref - timedelta(days=1)).strftime("%Y-W%W")
    else:
        semana_str = fecha_ref.strftime("%Y-W%W")

    semana_df = reg[reg["semana"] == semana_str].sort_values("fecha")
    if semana_df.empty:
        print(f"  [!] Sin datos para la semana {semana_str}")
        return

    dir_s = base / "reportes" / "semanales"
    dir_s.mkdir(parents=True, exist_ok=True)
    ruta_xlsx = str(dir_s / f"reporte_{semana_str.replace(':', '-')}.xlsx")
    reporte_semanal(cfg, semana_df, semana_str, ruta_xlsx)
    print(f"  → {ruta_xlsx}")


def procesar_mensual(cfg, reg, fecha_ref):
    base = cfg["_base"]
    reg = reg.copy()
    reg["fecha"] = pd.to_datetime(reg["fecha"], format='mixed')
    if fecha_ref.day == 1:
        mes_str = (fecha_ref - timedelta(days=1)).strftime("%Y-%m")
    else:
        mes_str = fecha_ref.strftime("%Y-%m")

    mes_df = reg[reg["mes"] == mes_str].sort_values("fecha")
    if mes_df.empty:
        print(f"  [!] Sin datos para el mes {mes_str}")
        return

    # DSM del último día del mes para LandXML
    ultimo_dia = pd.Timestamp(mes_df.iloc[-1]["fecha"]).date()
    dsm_ultimo = base / "vuelos" / ultimo_dia.isoformat() / "dsm.tif"
    baseline_p = ruta(cfg, "dem_baseline")
    eje_p      = ruta(cfg, "eje_via")

    # Heatmap mensual: ΔZ del último DSM del mes vs baseline
    arr_base, tf_ref, crs_ref, meta_ref = cargar_dem(str(baseline_p))
    arr_fin = alinear_dem(str(dsm_ultimo), tf_ref, crs_ref, arr_base.shape)
    dz_mes  = arr_fin - arr_base
    tmp_mes = str(base / "vuelos" / ultimo_dia.isoformat() / "dz_mes.tif")
    guardar_dz_raster(dz_mes, meta_ref, tmp_mes)

    gdf = gpd.read_file(str(eje_p)) if str(eje_p).endswith((".geojson", ".shp")) \
          else None
    if gdf is not None:
        if gdf.crs and gdf.crs != crs_ref: gdf = gdf.to_crs(crs_ref)
        eje  = gdf.geometry.iloc[0]
    else:
        eje = cargar_eje_dxf(str(eje_p), crs_ref)
    secs = crear_secciones(eje, cfg["espaciado_secciones"],
                           cfg["ancho_corredor"], cfg["prog_inicio_m"])

    dir_m = base / "reportes" / "mensuales"
    dir_m.mkdir(parents=True, exist_ok=True)
    heatmap_mes = str(dir_m / f"heatmap_{mes_str}.png")
    generar_heatmap(dz_mes, tf_ref, secs, eje,
                    f"ΔZ acumulado del mes {mes_str}",
                    heatmap_mes, cfg["dz_min_plot"], cfg["dz_max_plot"])

    ruta_xlsx = str(dir_m / f"reporte_{mes_str}.xlsx")
    reporte_mensual(cfg, mes_df, mes_str, heatmap_mes, ruta_xlsx)
    print(f"  → {ruta_xlsx}")

    # LandXML mensual
    ruta_xml = str(dir_m / f"superficies_civil3d_{mes_str}.xml")
    epsg = crs_ref.to_epsg() if crs_ref else None
    escribir_landxml(
        [("Baseline",       f"DEM inicial — {cfg['nombre']}", arr_base),
         (f"Fin_{mes_str}", f"DSM fin de mes {mes_str}",      arr_fin)],
        tf_ref, cfg["paso_muestreo_landxml"], ruta_xml, epsg)
    print(f"  → {ruta_xml}")


# ══════════════════════════════════════════════════════════════════════════════
# MAIN
# ══════════════════════════════════════════════════════════════════════════════
if __name__ == "__main__":
    parser = argparse.ArgumentParser(
        description="Pipeline de avance de obra vial — Dron + ODM")
    parser.add_argument("--fecha",   default=date.today().isoformat(),
                        help="Fecha del vuelo (YYYY-MM-DD). Default: hoy.")
    parser.add_argument("--semanal", action="store_true",
                        help="Forzar generación de reporte semanal.")
    parser.add_argument("--mensual", action="store_true",
                        help="Forzar generación de reporte mensual.")
    args = parser.parse_args()

    cfg   = cargar_config()
    fecha = date.fromisoformat(args.fecha)

    print("=" * 62)
    print(f"Pipeline de Avance — {cfg['nombre']}")
    print(f"Fecha: {fecha.isoformat()}")
    print("=" * 62)

    # Reporte diario (siempre)
    reg = procesar_dia(cfg, fecha)

    # Semanal: lunes automático o flag manual
    if fecha.weekday() == 0 or args.semanal:
        print("\n  Generando reporte semanal …")
        procesar_semanal(cfg, reg, fecha)

    # Mensual: día 1 automático o flag manual
    if fecha.day == 1 or args.mensual:
        print("\n  Generando reporte mensual …")
        procesar_mensual(cfg, reg, fecha)

    print("\n" + "=" * 62)
    print("  Listo.")
    print("=" * 62)
